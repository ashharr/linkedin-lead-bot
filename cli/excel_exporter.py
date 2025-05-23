import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import datetime
import logging
from typing import List, Dict, Any

logger = logging.getLogger(__name__)

class ExcelExporter:
    def generate_excel(self, leads_data: List[Dict[str, Any]], output_path: str) -> str:
        """Generates an Excel file from leads data and returns the path to the file."""
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "LinkedIn Leads"

        # Define headers
        headers = [
            "User Name", "Post Content", "Posted Timestamp", 
            "Profile URL", "Post URL", "Scraped At", "Search Query Ref"
        ]
        sheet.append(headers)

        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = openpyxl.styles.PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))

        for col_num, header_title in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Set column widths (approximate)
        sheet.column_dimensions['A'].width = 25  # User Name
        sheet.column_dimensions['B'].width = 60  # Post Content
        sheet.column_dimensions['C'].width = 20  # Posted Timestamp
        sheet.column_dimensions['D'].width = 40  # Profile URL
        sheet.column_dimensions['E'].width = 40  # Post URL
        sheet.column_dimensions['F'].width = 20  # Scraped At
        sheet.column_dimensions['G'].width = 25  # Search Query Ref

        # Populate data
        for row_num, lead in enumerate(leads_data, 2): # Start from row 2
            # Ensure datetime objects are naive before writing to Excel if they are timezone-aware
            # openpyxl does not handle timezone-aware datetime objects well by default.
            posted_ts = lead.get('posted_timestamp')
            if isinstance(posted_ts, datetime) and posted_ts.tzinfo is not None:
                posted_ts = posted_ts.replace(tzinfo=None)
            
            scraped_at_ts = lead.get('scraped_at')
            if isinstance(scraped_at_ts, datetime) and scraped_at_ts.tzinfo is not None:
                scraped_at_ts = scraped_at_ts.replace(tzinfo=None)

            row_data = [
                lead.get('user_name', 'N/A'),
                lead.get('post_content', 'N/A'),
                posted_ts,
                lead.get('profile_url', 'N/A'),
                lead.get('post_url', 'N/A'),
                scraped_at_ts,
                lead.get('search_query_ref', 'N/A')
            ]
            sheet.append(row_data)

            # Apply basic styling to data cells (borders, wrap text for content)
            for col_num in range(1, len(headers) + 1):
                cell = sheet.cell(row=row_num, column=col_num)
                cell.border = thin_border
                if col_num == 2: # Post Content column
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                else:
                    cell.alignment = Alignment(vertical='top')
                # Format date columns if they are datetime objects
                if col_num == 3 or col_num == 6: # Posted Timestamp & Scraped At
                    if isinstance(cell.value, datetime):
                        cell.number_format = 'YYYY-MM-DD HH:MM:SS'

        try:
            workbook.save(output_path)
            logger.info(f"Excel report generated successfully: {output_path}")
            return output_path
        except Exception as e:
            logger.error(f"Failed to save Excel report {output_path}: {e}")
            return ""

# Example Usage (for testing this module directly)
if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO)
    exporter = ExcelExporter()

    # Sample leads data (similar to what DataManager might provide)
    sample_leads_for_excel = [
        {
            "user_name": "Alice Example",
            "post_content": "This is Alice's first post. It is quite long and should wrap nicely in the Excel cell.",
            "posted_timestamp": datetime(2023, 1, 15, 10, 30, 0), # Naive datetime
            "profile_url": "http://linkedin.com/in/alicee",
            "post_url": "http://linkedin.com/feed/post/alice111",
            "scraped_at": datetime.now().replace(tzinfo=None), # Ensure naive for testing
            "search_query_ref": "query_A"
        },
        {
            "user_name": "Bob Sample",
            "post_content": "Bob's short post.",
            "posted_timestamp": "2023-01-16 12:00:00", # String date, will be written as string
            "profile_url": "http://linkedin.com/in/bobs",
            "post_url": "http://linkedin.com/feed/post/bob222",
            "scraped_at": "2023-01-17 09:00:00",
            "search_query_ref": "query_B"
        },
        {
            "user_name": "Charlie Test",
            # Missing some fields to test N/A handling
            "profile_url": "http://linkedin.com/in/charliet",
            "post_url": None, # Test None handling
            "scraped_at": datetime(2023,1,18,14,45,00) # Naive
        }
    ]

    output_file = "test_linkedin_leads.xlsx"
    generated_path = exporter.generate_excel(sample_leads_for_excel, output_file)

    if generated_path:
        print(f"Excel file generated: {generated_path}")
        # You can open this file to check its content and formatting.
    else:
        print("Failed to generate Excel file.")

    # Clean up
    import os
    if os.path.exists(output_file):
        # To prevent FileUsedByAnotherProcess error on Windows during cleanup in tests,
        # especially if the file was just opened by the user.
        try:
            os.remove(output_file)
            print(f"Cleaned up {output_file}")
        except PermissionError:
            print(f"Could not remove {output_file} due to PermissionError. Please close it if open.") 